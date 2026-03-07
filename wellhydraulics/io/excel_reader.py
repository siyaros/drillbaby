"""
Excel input file reader.

Reads the MATLAB-format Input*.xlsx workbook and returns a validated
ModelInput dataclass.  Matches all xlsread calls in hyds_mod.m.
"""

from __future__ import annotations

import numpy as np
import openpyxl
import logging
from pathlib import Path

from ..config.schema import (
    ModelInput, WellpathData, FormationData, CasingData, HoleData,
    TemperatureInput, KickData, DrillStringData, DrillStringSegment,
    SurfaceEquipment, FluidData, PVTCoefficients, PVTData,
    RealTimeStep, CvCurveData,
    MudBaseType, SaltType, InfluxType,
)

logger = logging.getLogger(__name__)


def _read_sheet(wb, name: str) -> list[list]:
    """Read a sheet into a list of rows (skipping header row)."""
    ws = wb[name]
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue  # skip header
        vals = [v if v is not None else 0 for v in row]
        rows.append(vals)
    return rows


def _float(v, default=0.0) -> float:
    try:
        return float(v)
    except (TypeError, ValueError):
        return default


def read_input_excel(filepath: str | Path) -> ModelInput:
    """Parse the MATLAB-format input Excel file.

    Parameters
    ----------
    filepath : path to Input*.xlsx

    Returns
    -------
    ModelInput with all validated data.
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)

    # ── Wellpath ─────────────────────────────────────────────────
    rows = _read_sheet(wb, "Wellpath")
    md_w = np.array([_float(r[0]) for r in rows])
    inc_w = np.array([_float(r[1]) for r in rows])
    wellpath = WellpathData(md=md_w, inclination=inc_w)

    # ── Formations ───────────────────────────────────────────────
    rows = _read_sheet(wb, "Formations")
    # Col layout: 0=Lithology 1=Name 2=Porosity 3=Perm 4=MD 5=TVD
    # 6=PPG 7=WBSG 8=FIT 9=FPG 10=LOT 11-13=HSGrad 14=ThCond
    fm_md = np.array([_float(r[4]) for r in rows])
    fm_ppg = np.array([_float(r[6]) for r in rows])
    fm_fpg = np.array([_float(r[9]) for r in rows])
    kf = _float(rows[0][14]) if len(rows[0]) > 14 else 0.98
    cpf = 0.22  # formation specific heat — not in this sheet, use default
    formations = FormationData(md=fm_md, ppg=fm_ppg, fpg=fm_fpg, k_thermal=kf, cp=cpf)

    # ── Casings ──────────────────────────────────────────────────
    rows = _read_sheet(wb, "Casings")
    # Col: 0=Type 1=OD 2=NW 3=ID 4=SD 5=HD
    if rows:
        c_od = np.array([_float(r[1]) for r in rows])
        c_hid = np.array([_float(r[3]) for r in rows])
        c_sd = np.array([_float(r[4]) for r in rows])
        c_hd = np.array([_float(r[5]) for r in rows])
    else:
        c_od = np.array([8.625])
        c_hid = np.array([8.0])
        c_sd = np.array([5000.0])
        c_hd = np.array([5000.0])
    casings = CasingData(od=c_od, hid=c_hid, sd=c_sd, hd=c_hd)

    # ── Hole ─────────────────────────────────────────────────────
    rows = _read_sheet(wb, "Hole")
    h_md = np.array([_float(r[0]) for r in rows])
    h_dia = np.array([_float(r[1]) for r in rows])
    hole = HoleData(md=h_md, diameter=h_dia)

    # ── Temperature ──────────────────────────────────────────────
    rows = _read_sheet(wb, "Temp")
    t_tvd = np.array([_float(r[0]) for r in rows])
    t_temp = np.array([_float(r[1]) for r in rows])
    temperature = TemperatureInput(tvd=t_tvd, temp=t_temp)

    # ── Kick ─────────────────────────────────────────────────────
    rows = _read_sheet(wb, "Kick")
    r0 = rows[0] if rows else [9950, 1000, 0.6, 2, 50]
    kick = KickData(
        md=_float(r0[0]), drain_radius=_float(r0[1]),
        gas_sg=_float(r0[2]), influx_type=InfluxType(int(_float(r0[3]))),
        volume=_float(r0[4]) if len(r0) > 4 else 0.0,
    )

    # ── Drill String ─────────────────────────────────────────────
    rows = _read_sheet(wb, "DS")
    segments = []
    ks_steel = _float(rows[0][13]) if len(rows[0]) > 13 else 24.85
    for r in rows:
        desc = str(r[0]) if r[0] else "pipe"
        tj_on = str(r[16]).strip() == '1' if len(r) > 16 and r[16] else False
        # Column layout (0-indexed): 0=Desc 1=Qty 2=Bottom 3=Top
        # 4=OD 5=ID 6=MaxD 7=ItemLen 8=TotalLen 9=AccLen
        # 10=NomWt 11=TotalWt 12=AccWt 13=ThCond 14=BSR 15=NonMag
        # 16=TJ-option 17=TJOD 18=TJID 19=TJL 20=NozzleType 21=NN 22=NI
        seg = DrillStringSegment(
            description=desc,
            od=_float(r[4]),   # OD, in
            pid=_float(r[5]),  # ID, in
            length=_float(r[7]),  # item length, ft
            total_length=_float(r[9]),  # accumulated length, ft
            tj_enabled=tj_on,
            tj_od=_float(r[17]) if len(r) > 17 else _float(r[4]),
            tj_id=_float(r[18]) if len(r) > 18 else _float(r[5]),
            tj_length=_float(r[19]) if len(r) > 19 else 1.0,
            k_thermal=ks_steel,
            num_nozzles=int(_float(r[21])) if len(r) > 21 else 0,
            nozzle_size=int(_float(r[22])) if len(r) > 22 else 0,
        )
        segments.append(seg)
    drillstring = DrillStringData(segments=segments, steel_conductivity=ks_steel)

    # ── Surface Equipment Inlet ──────────────────────────────────
    # Columns: 0=Type, 1=Length, 2=ID
    rows = _read_sheet(wb, "SEin")
    se_in_L = np.array([_float(r[1]) for r in rows]) if rows else np.array([])
    se_in_D = np.array([_float(r[2]) for r in rows]) if rows else np.array([])
    se_inlet = SurfaceEquipment(lengths=se_in_L, diameters=se_in_D)

    # ── Surface Equipment Outlet ─────────────────────────────────
    # Columns: 0=Type, 1=Length, 2=ID
    # Last row may be booster info: ['Booster=', depth, 'Lb[ft]']
    rows = _read_sheet(wb, "SEout")
    booster_depth = 0.0
    se_rows = rows
    if len(rows) >= 1:
        last = rows[-1]
        if isinstance(last[0], str) and 'booster' in str(last[0]).lower():
            booster_depth = _float(last[1])  # depth is in col 1
            se_rows = rows[:-1]

    se_out_L = np.array([_float(r[1]) for r in se_rows]) if se_rows else np.array([])
    se_out_D = np.array([_float(r[2]) for r in se_rows]) if se_rows else np.array([])
    se_outlet = SurfaceEquipment(lengths=se_out_L, diameters=se_out_D)

    # ── Fluids ───────────────────────────────────────────────────
    rows = _read_sheet(wb, "Fluids")
    fluid_list = []
    for r in rows:
        base_val = int(_float(r[1]))
        if base_val not in (1, 2, 3, 4):
            base_val = 4
        fl = FluidData(
            index=int(_float(r[0])),
            base=MudBaseType(base_val),
            temp_ref=_float(r[2]),
            pvt_enabled=bool(int(_float(r[3]))),
            f_base_pct=_float(r[4]),
            f_brine_pct=_float(r[5]),
            salt_type=SaltType(int(_float(r[6]))) if _float(r[6]) in (1, 2) else SaltType.CaCl2,
            salinity_pct=_float(r[7]),
            k_thermal=_float(r[8]),
            cp=_float(r[9]),
            tau_y=_float(r[10]),
            n=max(0.01, _float(r[11])),
            K=max(1e-10, _float(r[12])),
        )
        fluid_list.append(fl)

    # ── PVT ──────────────────────────────────────────────────────
    ws_pvt = wb["PVT"]
    pvt_rows = []
    for row in ws_pvt.iter_rows(min_row=2, values_only=True):
        pvt_rows.append([_float(v) for v in row[1:]])  # skip label column

    def _pvt_col(col_idx):
        return PVTCoefficients(
            a1=pvt_rows[0][col_idx], b1=pvt_rows[1][col_idx],
            c1=pvt_rows[2][col_idx], a2=pvt_rows[3][col_idx],
            b2=pvt_rows[4][col_idx], c2=pvt_rows[5][col_idx],
        )

    pvt_data = PVTData(
        base_coeffs={
            MudBaseType.MOBM: _pvt_col(0),
            MudBaseType.SBM: _pvt_col(1),
            MudBaseType.OBM: _pvt_col(2),
            MudBaseType.WBM: _pvt_col(3),
        },
        water_coeffs=_pvt_col(3),  # WBM = pure water at s=0
        cacl2_s1=_pvt_col(4),
        cacl2_s2=_pvt_col(5),
        nacl_s1=_pvt_col(6),
        nacl_s2=_pvt_col(7),
    )

    # ── RealTime ─────────────────────────────────────────────────
    rows = _read_sheet(wb, "RealTime")
    rt_steps = []
    for r in rows:
        step = RealTimeStep(
            time=_float(r[0]),
            Q=_float(r[1]),
            SBP=_float(r[2]),
            BHPC_enabled=bool(int(_float(r[3]))),
            BHP_setpoint=_float(r[4]),
            gain=_float(r[5]),
            Q_booster=_float(r[6]),
            mud_index=int(_float(r[7])),
            density_ref=_float(r[9]),
            T_inlet=_float(r[11]),
            RPM=_float(r[12]),
            bit_depth=_float(r[13]),
            T_calib_enabled=bool(int(_float(r[14]))) if len(r) > 14 else False,
            BHT_measured=_float(r[15]) if len(r) > 15 else 100.0,
            T_out_measured=_float(r[16]) if len(r) > 16 else 90.0,
            Q_out=_float(r[17]) if len(r) > 17 else _float(r[1]),
        )
        rt_steps.append(step)

    # ── Cv-curve ─────────────────────────────────────────────────
    cv_curve = None
    if "Cv-curve" in wb.sheetnames:
        rows = _read_sheet(wb, "Cv-curve")
        if rows:
            cv_vals = np.array([_float(r[0]) for r in rows])
            op_vals = np.array([_float(r[1]) for r in rows])
            cv_curve = CvCurveData(cv=cv_vals, op=op_vals)

    wb.close()

    return ModelInput(
        wellpath=wellpath, formations=formations,
        casings=casings, hole=hole, temperature=temperature,
        kick=kick, drillstring=drillstring,
        se_inlet=se_inlet, se_outlet=se_outlet,
        booster_depth=booster_depth,
        fluids=fluid_list, pvt=pvt_data,
        realtime=rt_steps, cv_curve=cv_curve,
    )
